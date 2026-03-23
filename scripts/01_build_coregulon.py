#!/usr/bin/env python3
"""
01_build_coregulon.py
=====================
ZORC — Zip-code Of RNAs that Condense
Phase 1: Build the coregulon gene list for ML training.

Intersects:
  - T-RIP RNA enriched/depleted lists (Liu et al. 2024, Plant Cell, Suppl. Data Set 2)
  - APEAL proteomics (Liu et al. 2023, EMBO J, Source File 1)

POSITIVE class (1): RNA enriched in P-bodies (NS or HS) AND corresponding
    protein detected with log2FC >= threshold in ANY APEAL method/condition
    (AP_NS, AP_HS, PDL_NS, PDL_HS — union, not intersection).

NEGATIVE class (0): RNA depleted from P-bodies (NS or HS), regardless of
    protein detection status.

NOTE on isoform vs canonical protein:
    The gene ID (AGI code) from APEAL is used ONLY as a matching key to confirm
    P-body protein presence. Sequence features in downstream steps (P2–P7) will
    use the SPECIFIC isoform sequence identified in T-RIP, NOT the canonical
    protein. This is a deliberate methodological choice documented here and in
    config.yaml.

Usage:
    python 01_build_coregulon.py --config config/zorc_config.yaml

Output:
    data/processed/01_zorc_coregulon_list.csv
    logs/01_coregulon_build_report.txt

Author: José Moya-Cuevas, MoschouLab / IMBB-FORTH / ERC PLANTEX
Project: ZORC — github.com/MoschouLab/ZORC
License: MIT
"""

import argparse
import os
import sys
import textwrap
from datetime import datetime
from pathlib import Path

import pandas as pd
import yaml


# =============================================================================
# CONFIG LOADING
# =============================================================================

def load_config(config_path: str) -> dict:
    with open(config_path) as f:
        cfg = yaml.safe_load(f)
    return cfg


def resolve_path(p: str) -> Path:
    return Path(os.path.expanduser(p))


# =============================================================================
# DATA LOADING — APEAL PROTEOMICS (EMBO J)
# =============================================================================

def load_apeal_proteins(cfg: dict) -> pd.DataFrame:
    """
    Load APEAL dataset from EMBO J Source File 1.

    Column mapping (0-indexed, from header row 2):
      col 0:  protein (AGI gene ID)
      col 1:  description
      col 32: Log2(FC) AP NS   (DCP1/GFP)
      col 33: Log2(FC) AP HS
      col 34: Log2(FC) PDL NS
      col 35: Log2(FC) PDL HS
    """
    path = resolve_path(
        os.path.join(cfg['project_dir'], cfg['input_files']['apeal_excel'])
    )
    sheet = cfg['input_files']['apeal_sheet']

    print(f"  Loading APEAL data from: {path}")
    print(f"  Sheet: '{sheet}'")

    from openpyxl import load_workbook
    wb = load_workbook(str(path), read_only=True)
    ws = wb[sheet]

    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3:
            continue  # skip 2 title rows + 1 section header
        if row[0] is None:
            continue
        gid = str(row[0]).strip()
        if not (gid.startswith('AT') and len(gid) >= 9 and gid[2].isdigit()):
            continue
        rows.append({
            'geneID':          gid,
            'description_prot': str(row[1]).strip() if row[1] else '',
            'log2FC_AP_NS':    row[32],
            'log2FC_AP_HS':    row[33],
            'log2FC_PDL_NS':   row[34],
            'log2FC_PDL_HS':   row[35],
        })

    df = pd.DataFrame(rows)
    for col in ['log2FC_AP_NS', 'log2FC_AP_HS', 'log2FC_PDL_NS', 'log2FC_PDL_HS']:
        df[col] = pd.to_numeric(df[col], errors='coerce')

    print(f"  → {len(df):,} proteins loaded from APEAL dataset")
    return df


# =============================================================================
# DATA LOADING — T-RIP RNA LISTS (Plant Cell, Suppl. Data Set 2)
# =============================================================================

def load_trip_rna(cfg: dict) -> pd.DataFrame:
    """
    Load T-RIP RNA enriched/depleted lists from Suppl. Data Set 2.

    Sheet layout (4 parallel sections, separated by empty columns):
      Cols  0-3:  PBs-enriched NS  (geneID, Description, log2FC_NS, log2FC_HS)
      Cols  5-8:  PBs-enriched HS
      Cols 10-13: PBs-depleted NS
      Cols 15-18: PBs-depleted HS

    Returns a per-gene DataFrame with flags for each list membership and
    the strongest log2FC signal per condition.
    """
    path = resolve_path(
        os.path.join(cfg['project_dir'], cfg['input_files']['trip_excel'])
    )
    sheet  = cfg['input_files']['trip_sheet']
    skip_n = cfg['input_files']['trip_header_rows']

    print(f"  Loading T-RIP RNA data from: {path}")
    print(f"  Sheet: '{sheet}', skipping {skip_n} header rows")

    from openpyxl import load_workbook
    wb = load_workbook(str(path), read_only=True)
    ws = wb[sheet]

    enr_ns, enr_hs, dep_ns, dep_hs = [], [], [], []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < skip_n:
            continue

        def _at(val):
            if val is None:
                return None
            s = str(val).strip()
            return s if (s.startswith('AT') and len(s) >= 9 and s[2].isdigit()) else None

        # Enriched NS (cols 0-3)
        if _at(row[0]):
            enr_ns.append({
                'geneID':         row[0],
                'description_rna': str(row[1]).strip() if row[1] else '',
                'rna_log2FC_NS':  row[2],
                'rna_log2FC_HS':  row[3],
            })
        # Enriched HS (cols 5-8)
        if _at(row[5]):
            enr_hs.append({
                'geneID':         row[5],
                'description_rna': str(row[6]).strip() if row[6] else '',
                'rna_log2FC_NS':  row[7],
                'rna_log2FC_HS':  row[8],
            })
        # Depleted NS (cols 10-13)
        if _at(row[10]):
            dep_ns.append({
                'geneID':         row[10],
                'description_rna': str(row[11]).strip() if row[11] else '',
                'rna_log2FC_NS':  row[12],
                'rna_log2FC_HS':  row[13],
            })
        # Depleted HS (cols 15-18)
        if _at(row[15]):
            dep_hs.append({
                'geneID':         row[15],
                'description_rna': str(row[16]).strip() if row[16] else '',
                'rna_log2FC_NS':  row[17],
                'rna_log2FC_HS':  row[18],
            })

    # Build per-list dataframes with membership flags
    def _make_df(records, flag_col):
        df = pd.DataFrame(records)
        df[flag_col] = True
        return df[['geneID', flag_col, 'description_rna', 'rna_log2FC_NS', 'rna_log2FC_HS']]

    df_enr_ns = _make_df(enr_ns, 'enriched_NS_RNA')
    df_enr_hs = _make_df(enr_hs, 'enriched_HS_RNA')
    df_dep_ns = _make_df(dep_ns, 'depleted_NS_RNA')
    df_dep_hs = _make_df(dep_hs, 'depleted_HS_RNA')

    # Merge all four lists into a per-gene master table
    all_pieces = pd.concat([df_enr_ns, df_enr_hs, df_dep_ns, df_dep_hs], ignore_index=True)
    all_pieces['geneID'] = all_pieces['geneID'].astype(str).str.strip()

    for col in ['rna_log2FC_NS', 'rna_log2FC_HS']:
        all_pieces[col] = pd.to_numeric(all_pieces[col], errors='coerce')

    # Aggregate: one row per gene
    flag_cols = ['enriched_NS_RNA', 'enriched_HS_RNA', 'depleted_NS_RNA', 'depleted_HS_RNA']
    agg = (
        all_pieces.groupby('geneID', sort=False)
        .agg(
            description_rna  = ('description_rna', 'first'),
            rna_log2FC_NS    = ('rna_log2FC_NS', 'first'),
            rna_log2FC_HS    = ('rna_log2FC_HS', 'first'),
            enriched_NS_RNA  = ('enriched_NS_RNA', lambda x: True in x.values),
            enriched_HS_RNA  = ('enriched_HS_RNA', lambda x: True in x.values),
            depleted_NS_RNA  = ('depleted_NS_RNA', lambda x: True in x.values),
            depleted_HS_RNA  = ('depleted_HS_RNA', lambda x: True in x.values),
        )
        .reset_index()
    )
    for fc in flag_cols:
        agg[fc] = agg[fc].fillna(False)

    n_enr = (agg['enriched_NS_RNA'] | agg['enriched_HS_RNA']).sum()
    n_dep = (agg['depleted_NS_RNA'] | agg['depleted_HS_RNA']).sum()
    print(f"  → {len(agg):,} unique genes in RNA lists")
    print(f"     Enriched (NS∪HS): {n_enr:,}")
    print(f"     Depleted (NS∪HS): {n_dep:,}")

    return agg


# =============================================================================
# COREGULON CONSTRUCTION
# =============================================================================

def build_coregulon(
    df_prot: pd.DataFrame,
    df_rna:  pd.DataFrame,
    cfg:     dict,
) -> pd.DataFrame:
    """
    Intersect RNA and protein data to build the ZORC training set.

    POSITIVE (class=1):
        RNA enriched in P-bodies (NS or HS)
        AND protein log2FC >= threshold in ANY of {AP_NS, AP_HS, PDL_NS, PDL_HS}

    NEGATIVE (class=0):
        RNA depleted from P-bodies (NS or HS)
        (protein detection status irrelevant for negatives)

    Conflict genes (enriched AND depleted): resolved per config.
    """
    thr  = cfg['coregulon']['protein_enrichment_threshold']
    conf = cfg['coregulon']['conflict_resolution']
    pos_label = cfg['classes']['positive_label']
    neg_label = cfg['classes']['negative_label']

    print(f"\n  Protein enrichment threshold: log2FC >= {thr}")
    print(f"  Applied to: ANY of (AP_NS, AP_HS, PDL_NS, PDL_HS)")
    print(f"  Conflict resolution strategy: '{conf}'")

    # --- Flag protein enrichment ---------------------------------------------
    prot_cols = ['log2FC_AP_NS', 'log2FC_AP_HS', 'log2FC_PDL_NS', 'log2FC_PDL_HS']
    df_prot = df_prot.copy()
    df_prot['protein_enriched'] = df_prot[prot_cols].ge(thr).any(axis=1)

    enriched_protein_ids = set(df_prot.loc[df_prot['protein_enriched'], 'geneID'])
    print(f"  Proteins with log2FC >= {thr} in any method/condition: {len(enriched_protein_ids):,}")

    # --- RNA gene sets -------------------------------------------------------
    df_rna = df_rna.copy()
    rna_enriched_mask = df_rna['enriched_NS_RNA'] | df_rna['enriched_HS_RNA']
    rna_depleted_mask = df_rna['depleted_NS_RNA'] | df_rna['depleted_HS_RNA']

    rna_enriched_ids = set(df_rna.loc[rna_enriched_mask, 'geneID'])
    rna_depleted_ids = set(df_rna.loc[rna_depleted_mask, 'geneID'])

    # --- Raw positive and negative sets -------------------------------------
    raw_positives = rna_enriched_ids & enriched_protein_ids
    raw_negatives = rna_depleted_ids  # no protein filter for negatives

    conflict_ids = raw_positives & raw_negatives
    print(f"\n  Raw positives (RNA enriched ∩ protein ≥ {thr}): {len(raw_positives):,}")
    print(f"  Raw negatives (RNA depleted): {len(raw_negatives):,}")
    print(f"  Conflict genes (in both): {len(conflict_ids):,}")

    # --- Resolve conflicts ---------------------------------------------------
    resolved_positives = raw_positives.copy()
    resolved_negatives = raw_negatives.copy()
    conflict_disposition = {}

    if conflict_ids:
        for gid in conflict_ids:
            row = df_rna[df_rna['geneID'] == gid]
            if row.empty:
                conflict_disposition[gid] = 'dropped_no_data'
                resolved_positives.discard(gid)
                resolved_negatives.discard(gid)
                continue

            fc_ns = float(row['rna_log2FC_NS'].iloc[0]) if pd.notna(row['rna_log2FC_NS'].iloc[0]) else 0.0
            fc_hs = float(row['rna_log2FC_HS'].iloc[0]) if pd.notna(row['rna_log2FC_HS'].iloc[0]) else 0.0

            if conf == 'enriched':
                resolved_negatives.discard(gid)
                conflict_disposition[gid] = 'kept_as_positive'
            elif conf == 'depleted':
                resolved_positives.discard(gid)
                conflict_disposition[gid] = 'kept_as_negative'
            elif conf == 'drop':
                resolved_positives.discard(gid)
                resolved_negatives.discard(gid)
                conflict_disposition[gid] = 'dropped'
            elif conf == 'strongest':
                # Use the list where |log2FC| is strongest
                # Enriched signal: use the positive fc (should be > 0)
                # Depleted signal: use negative fc (should be < 0)
                enr_signal = max(abs(fc_ns) if fc_ns > 0 else 0,
                                 abs(fc_hs) if fc_hs > 0 else 0)
                dep_signal = max(abs(fc_ns) if fc_ns < 0 else 0,
                                 abs(fc_hs) if fc_hs < 0 else 0)
                if enr_signal >= dep_signal:
                    resolved_negatives.discard(gid)
                    conflict_disposition[gid] = f'kept_as_positive (enr={enr_signal:.2f} > dep={dep_signal:.2f})'
                else:
                    resolved_positives.discard(gid)
                    conflict_disposition[gid] = f'kept_as_negative (dep={dep_signal:.2f} > enr={enr_signal:.2f})'
            else:
                raise ValueError(f"Unknown conflict_resolution strategy: '{conf}'")

        print(f"  Conflict resolutions:")
        for gid, disp in conflict_disposition.items():
            print(f"    {gid}: {disp}")

    print(f"\n  Final positives: {len(resolved_positives):,}")
    print(f"  Final negatives: {len(resolved_negatives):,}")

    # --- Assemble output dataframe -------------------------------------------
    # Merge RNA + protein data for all genes in final set
    all_genes = sorted(resolved_positives | resolved_negatives)

    df_out = df_rna[df_rna['geneID'].isin(all_genes)].copy()
    df_out = df_out.merge(
        df_prot[['geneID', 'description_prot'] + prot_cols + ['protein_enriched']],
        on='geneID', how='left'
    )

    # Assign class labels
    df_out['class'] = df_out['geneID'].apply(
        lambda g: pos_label if g in resolved_positives else neg_label
    )

    # Condition column: which condition drives the classification (for positives)
    def _condition(row):
        if row['class'] == neg_label:
            return 'depleted'
        enr_ns = row.get('enriched_NS_RNA', False)
        enr_hs = row.get('enriched_HS_RNA', False)
        if enr_ns and enr_hs:
            return 'both'
        elif enr_ns:
            return 'NS'
        else:
            return 'HS'

    df_out['condition'] = df_out.apply(_condition, axis=1)

    # strongest_rna_log2fc: absolute value of the strongest RNA FC signal
    df_out['strongest_rna_log2fc'] = df_out.apply(
        lambda r: max(
            abs(r['rna_log2FC_NS']) if pd.notna(r['rna_log2FC_NS']) else 0,
            abs(r['rna_log2FC_HS']) if pd.notna(r['rna_log2FC_HS']) else 0
        ),
        axis=1
    )

    # conflict flag
    df_out['conflict_resolved'] = df_out['geneID'].isin(conflict_ids)

    # Final column order
    col_order = [
        'geneID', 'description_rna', 'description_prot',
        'class', 'condition',
        'enriched_NS_RNA', 'enriched_HS_RNA',
        'depleted_NS_RNA', 'depleted_HS_RNA',
        'rna_log2FC_NS', 'rna_log2FC_HS', 'strongest_rna_log2fc',
        'log2FC_AP_NS', 'log2FC_AP_HS',
        'log2FC_PDL_NS', 'log2FC_PDL_HS',
        'protein_enriched', 'conflict_resolved',
    ]
    df_out = df_out[[c for c in col_order if c in df_out.columns]]
    df_out = df_out.sort_values(['class', 'strongest_rna_log2fc'], ascending=[False, False])
    df_out = df_out.reset_index(drop=True)

    return df_out, conflict_disposition


# =============================================================================
# AUDIT REPORT
# =============================================================================

def write_report(
    df_out: pd.DataFrame,
    conflict_disposition: dict,
    cfg: dict,
    output_path: str,
):
    pos_label = cfg['classes']['positive_label']
    neg_label = cfg['classes']['negative_label']
    thr = cfg['coregulon']['protein_enrichment_threshold']

    n_pos = (df_out['class'] == pos_label).sum()
    n_neg = (df_out['class'] == neg_label).sum()
    n_conflict = df_out['conflict_resolved'].sum()
    ratio = n_neg / n_pos if n_pos > 0 else float('inf')

    condition_counts = df_out[df_out['class'] == pos_label]['condition'].value_counts()

    report = textwrap.dedent(f"""
    ================================================================================
    ZORC — Coregulon Build Report
    Script: 01_build_coregulon.py
    Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
    Config: {cfg.get('_config_path', 'N/A')}
    ================================================================================

    PARAMETERS
    ----------
    Protein enrichment threshold : log2FC >= {thr} in ANY of (AP_NS, AP_HS, PDL_NS, PDL_HS)
    Conflict resolution strategy : {cfg['coregulon']['conflict_resolution']}
    RNA min abs log2FC           : {cfg['coregulon']['rna_min_abs_log2fc']}

    INPUT SOURCES
    -------------
    APEAL proteomics : {cfg['input_files']['apeal_excel']}
      Sheet          : {cfg['input_files']['apeal_sheet']}
      Reference      : Liu C, Mentzelopoulou A et al. EMBO J. 2023;42(9):e111885.
                       doi:10.15252/embj.2022111885

    T-RIP RNA lists  : {cfg['input_files']['trip_excel']}
      Sheet          : {cfg['input_files']['trip_sheet']}
      Reference      : Liu et al. 2024, Plant Cell. Suppl. Data Set 2.
                       doi:10.1093/plcell/koae015

    RESULTS SUMMARY
    ---------------
    Total genes in coregulon : {len(df_out):,}
    Positives (class=1)      : {n_pos:,}
      - Enriched NS only     : {condition_counts.get('NS', 0):,}
      - Enriched HS only     : {condition_counts.get('HS', 0):,}
      - Enriched in both     : {condition_counts.get('both', 0):,}
    Negatives (class=0)      : {n_neg:,}
    Class ratio (neg:pos)    : 1:{ratio:.2f}
    Conflicts resolved       : {n_conflict:,}

    BIOLOGICAL INTERPRETATION
    -------------------------
    Positives represent mRNAs that are:
      (1) Enriched in DCP1-TurboID pulldowns (P-body co-enrichment in T-RIP)
      (2) Whose gene product is also detected with log2FC >= {thr} in at least
          one APEAL condition (AP or PDL, NS or HS).
    This "coregulon" criterion captures cogulate mRNA-protein pairs that are
    jointly present in P-bodies under at least one stress condition.

    Negatives represent mRNAs actively depleted from P-bodies, serving as
    true negative examples for the ZIP-CODE prediction model.

    IMPORTANT — Isoform vs canonical protein:
      The AGI code from APEAL is used ONLY as a matching key. Downstream
      sequence features (P2–P7) will use the specific ISOFORM sequence
      identified in T-RIP, not the canonical protein sequence. See config.yaml.

    CONFLICT RESOLUTIONS ({n_conflict} genes)
    {'-'*40}
    """)

    if conflict_disposition:
        for gid, disp in conflict_disposition.items():
            report += f"    {gid}: {disp}\n"
    else:
        report += "    None\n"

    report += textwrap.dedent(f"""
    OUTPUT
    ------
    Coregulon list: {cfg['outputs']['coregulon_list']}

    Columns:
      geneID              - AGI locus identifier (TAIR10)
      description_rna     - gene description from T-RIP dataset
      description_prot    - protein description from APEAL dataset
      class               - 1=positive (enriched), 0=negative (depleted)
      condition           - NS / HS / both / depleted
      enriched_NS_RNA     - bool: in T-RIP enriched NS list
      enriched_HS_RNA     - bool: in T-RIP enriched HS list
      depleted_NS_RNA     - bool: in T-RIP depleted NS list
      depleted_HS_RNA     - bool: in T-RIP depleted HS list
      rna_log2FC_NS       - log2FC DCP1/GFP in NS (T-RIP)
      rna_log2FC_HS       - log2FC DCP1/GFP in HS (T-RIP)
      strongest_rna_log2fc- max(|log2FC_NS|, |log2FC_HS|)
      log2FC_AP_NS        - log2FC DCP1/GFP in AP NS (APEAL)
      log2FC_AP_HS        - log2FC DCP1/GFP in AP HS (APEAL)
      log2FC_PDL_NS       - log2FC DCP1/GFP in PDL NS (APEAL)
      log2FC_PDL_HS       - log2FC DCP1/GFP in PDL HS (APEAL)
      protein_enriched    - bool: protein_log2FC >= {thr} in any APEAL condition
      conflict_resolved   - bool: gene appeared in both enriched and depleted lists

    ================================================================================
    END OF REPORT
    ================================================================================
    """)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, 'w') as f:
        f.write(report)

    print(f"\n  Report written to: {output_path}")
    return report


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="ZORC Phase 1 — Build coregulon gene list for ML training.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=textwrap.dedent("""
            Examples:
              python 01_build_coregulon.py --config config/zorc_config.yaml
              python 01_build_coregulon.py --config config/zorc_config.yaml --dry-run
        """)
    )
    parser.add_argument(
        '--config', required=True,
        help='Path to zorc_config.yaml'
    )
    parser.add_argument(
        '--dry-run', action='store_true',
        help='Load and parse data but do not write output files'
    )
    parser.add_argument(
        '--protein-threshold', type=float, default=None,
        help='Override protein_enrichment_threshold from config'
    )
    args = parser.parse_args()

    cfg = load_config(args.config)
    cfg['_config_path'] = args.config

    # Override threshold from CLI if provided
    if args.protein_threshold is not None:
        print(f"  [CLI override] protein_enrichment_threshold: "
              f"{cfg['coregulon']['protein_enrichment_threshold']} → {args.protein_threshold}")
        cfg['coregulon']['protein_enrichment_threshold'] = args.protein_threshold

    print("=" * 70)
    print("ZORC — Phase 1: Build Coregulon List")
    print(f"Config: {args.config}")
    print(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)

    # --- Load data -----------------------------------------------------------
    print("\n[1/4] Loading APEAL proteomics (EMBO J 2023)...")
    df_prot = load_apeal_proteins(cfg)

    print("\n[2/4] Loading T-RIP RNA lists (Plant Cell 2024)...")
    df_rna = load_trip_rna(cfg)

    # --- Build coregulon -----------------------------------------------------
    print("\n[3/4] Building coregulon intersection...")
    df_coregulon, conflict_disp = build_coregulon(df_prot, df_rna, cfg)

    # --- Output summary ------------------------------------------------------
    pos_label = cfg['classes']['positive_label']
    neg_label = cfg['classes']['negative_label']
    n_pos = (df_coregulon['class'] == pos_label).sum()
    n_neg = (df_coregulon['class'] == neg_label).sum()

    print(f"\n{'='*70}")
    print(f"COREGULON SUMMARY")
    print(f"{'='*70}")
    print(f"  Total genes : {len(df_coregulon):,}")
    print(f"  Positives   : {n_pos:,}  (class=1, RNA enriched + protein ≥ {cfg['coregulon']['protein_enrichment_threshold']})")
    print(f"  Negatives   : {n_neg:,}  (class=0, RNA depleted)")
    print(f"  Ratio       : 1:{n_neg/n_pos:.2f}  ({'balanced ✓' if abs(n_neg/n_pos - 1) < 0.5 else 'imbalanced'})")
    print(f"{'='*70}")

    print("\nTop 10 positives (by strongest RNA log2FC):")
    top10 = df_coregulon[df_coregulon['class'] == pos_label].head(10)
    for _, row in top10.iterrows():
        print(f"  {row['geneID']:<14} log2FC_max={row['strongest_rna_log2fc']:.2f}"
              f"  cond={row['condition']:<5}  prot_AP_NS={row.get('log2FC_AP_NS', float('nan')):.2f}")

    # --- Write outputs -------------------------------------------------------
    print("\n[4/4] Writing outputs...")

    if not args.dry_run:
        out_csv = resolve_path(
            os.path.join(cfg['project_dir'], cfg['outputs']['coregulon_list'])
        )
        out_csv.parent.mkdir(parents=True, exist_ok=True)
        df_coregulon.to_csv(out_csv, index=False)
        print(f"  Coregulon list: {out_csv}")
        print(f"  Rows: {len(df_coregulon):,}  |  Columns: {len(df_coregulon.columns)}")

        report_path = resolve_path(
            os.path.join(cfg['project_dir'], cfg['outputs']['coregulon_report'])
        )
        report_text = write_report(df_coregulon, conflict_disp, cfg, str(report_path))
        print(report_text)
    else:
        print("  [DRY RUN] — no files written")

    print("\n✓ Phase 1 complete.")
    return 0


if __name__ == '__main__':
    sys.exit(main())
